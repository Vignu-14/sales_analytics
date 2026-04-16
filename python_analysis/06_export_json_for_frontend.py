"""Export frontend-ready JavaScript data and create the Excel dashboard workbook."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
try:
    from sklearn.linear_model import LinearRegression
    from sklearn.metrics import mean_squared_error, r2_score
except ImportError:  # pragma: no cover - lightweight fallback
    class LinearRegression:  # type: ignore[override]
        """Fallback linear regression using NumPy polyfit."""

        def fit(self, x_values, y_values):
            x_array = np.asarray(x_values).reshape(-1)
            y_array = np.asarray(y_values)
            slope, intercept = np.polyfit(x_array, y_array, 1)
            self.coef_ = np.array([slope])
            self.intercept_ = intercept
            return self

        def predict(self, x_values):
            x_array = np.asarray(x_values).reshape(-1)
            return (self.coef_[0] * x_array) + self.intercept_

    def mean_squared_error(y_true, y_pred, squared=True):
        error = np.mean((np.asarray(y_true) - np.asarray(y_pred)) ** 2)
        return error if squared else np.sqrt(error)

    def r2_score(y_true, y_pred):
        y_true = np.asarray(y_true)
        y_pred = np.asarray(y_pred)
        ss_res = np.sum((y_true - y_pred) ** 2)
        ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
        return 1 - (ss_res / ss_tot if ss_tot else 0)


BASE_DIR = Path(r"C:\anil")
INPUT_FILE = BASE_DIR / "cleaned_sales_data.csv"
DATA_JS_FILE = BASE_DIR / "frontend" / "js" / "data.js"
EXCEL_FILE = BASE_DIR / "excel_dashboard" / "Sales_Dashboard.xlsx"

PRIMARY = "1B2A4A"
WHITE = "FFFFFF"
LIGHT_ROW = "F7F9FC"


def configure_console() -> None:
    """Enable UTF-8 console output where supported."""
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass


def round_frame(dataframe: pd.DataFrame, digits: int = 2) -> pd.DataFrame:
    """Round all numeric columns for clean JSON export."""
    rounded = dataframe.copy()
    for column in rounded.select_dtypes(include=["number"]).columns:
        rounded[column] = rounded[column].round(digits)
    return rounded


def build_forecast(dataframe: pd.DataFrame) -> tuple[list[dict], dict[str, float]]:
    """Build forecast series and model metrics for the dashboard."""
    monthly = (
        dataframe.groupby("Year_Month")["Sales_Amount"]
        .sum()
        .reset_index()
        .sort_values("Year_Month")
    )
    monthly["Month_Date"] = pd.to_datetime(monthly["Year_Month"] + "-01")
    monthly["Time_Index"] = np.arange(1, len(monthly) + 1)

    model = LinearRegression()
    x_values = monthly[["Time_Index"]]
    y_values = monthly["Sales_Amount"]
    model.fit(x_values, y_values)
    monthly["Predicted"] = model.predict(x_values)

    future_index = np.arange(len(monthly) + 1, len(monthly) + 7)
    future_months = pd.date_range(monthly["Month_Date"].max() + pd.offsets.MonthBegin(1), periods=6, freq="MS")
    future_predictions = model.predict(future_index.reshape(-1, 1))

    forecast_records: list[dict] = []
    for _, row in monthly.iterrows():
        forecast_records.append(
            {
                "month": row["Month_Date"].strftime("%Y-%m"),
                "actual": round(float(row["Sales_Amount"]), 2),
                "predicted": round(float(row["Predicted"]), 2),
            }
        )

    for month_value, predicted_value in zip(future_months, future_predictions):
        forecast_records.append(
            {
                "month": month_value.strftime("%Y-%m"),
                "actual": None,
                "predicted": round(float(predicted_value), 2),
            }
        )

    metrics = {
        "r2": round(float(r2_score(y_values, monthly["Predicted"])), 4),
        "rmse": round(float(mean_squared_error(y_values, monthly["Predicted"], squared=False)), 2),
        "slope": round(float(model.coef_[0]), 2),
        "intercept": round(float(model.intercept_), 2),
        "confidenceBandPct": 15.0,
    }
    return forecast_records, metrics


def build_dashboard_payload(dataframe: pd.DataFrame) -> dict:
    """Generate the dashboard data structure used by the frontend."""
    total_sales = dataframe["Sales_Amount"].sum()
    total_profit = dataframe["Profit"].sum()
    total_orders = dataframe["Order ID"].nunique()
    avg_order_value = total_sales / total_orders if total_orders else 0
    profit_margin = (total_profit / total_sales * 100) if total_sales else 0

    monthly_sales = round_frame(
        dataframe.groupby("Year_Month")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"))
        .reset_index()
        .rename(columns={"Year_Month": "month"})
        .sort_values("month")
    )
    region_data = round_frame(
        dataframe.groupby("Region")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"), orders=("Order ID", "nunique"))
        .reset_index()
        .rename(columns={"Region": "region"})
        .sort_values("sales", ascending=False)
    )
    category_data = round_frame(
        dataframe.groupby("Product_Category")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"))
        .reset_index()
        .rename(columns={"Product_Category": "category"})
        .sort_values("sales", ascending=False)
    )
    subcategory_data = round_frame(
        dataframe.groupby("Sub_Category")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"))
        .reset_index()
        .rename(columns={"Sub_Category": "subCategory"})
        .sort_values("sales", ascending=False)
        .head(15)
    )
    product_summary = (
        dataframe.groupby("Product Name")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"))
        .reset_index()
        .rename(columns={"Product Name": "product"})
    )
    top_products = round_frame(product_summary.sort_values("sales", ascending=False).head(10))
    bottom_products = round_frame(product_summary.sort_values("profit", ascending=True).head(10))
    top_customers = round_frame(
        dataframe.groupby("Customer Name")
        .agg(sales=("Sales_Amount", "sum"), orders=("Order ID", "nunique"))
        .reset_index()
        .rename(columns={"Customer Name": "customer"})
        .sort_values("sales", ascending=False)
        .head(10)
    )
    top_states = round_frame(
        dataframe.groupby("State")
        .agg(sales=("Sales_Amount", "sum"))
        .reset_index()
        .rename(columns={"State": "state"})
        .sort_values("sales", ascending=False)
        .head(10)
    )
    city_data = round_frame(
        dataframe.groupby(["City", "State", "Region"])
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"), orders=("Order ID", "nunique"))
        .reset_index()
        .rename(columns={"City": "city", "State": "state", "Region": "region"})
        .sort_values("sales", ascending=False)
        .head(25)
    )
    segment_data = round_frame(
        dataframe.groupby("Segment")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"), orders=("Order ID", "nunique"))
        .reset_index()
        .rename(columns={"Segment": "segment"})
        .sort_values("sales", ascending=False)
    )
    payment_data = round_frame(
        dataframe.groupby("Payment_Method")
        .agg(sales=("Sales_Amount", "sum"), count=("Order ID", "nunique"))
        .reset_index()
        .rename(columns={"Payment_Method": "method"})
        .sort_values("sales", ascending=False)
    )
    channel_data = round_frame(
        dataframe.groupby("Sales_Channel")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"), count=("Order ID", "nunique"))
        .reset_index()
        .rename(columns={"Sales_Channel": "channel"})
        .sort_values("sales", ascending=False)
    )
    ship_mode_data = round_frame(
        dataframe.groupby("Ship Mode")
        .agg(orders=("Order ID", "nunique"), avgDays=("Shipping_Days", "mean"))
        .reset_index()
        .rename(columns={"Ship Mode": "mode"})
        .sort_values("orders", ascending=False)
    )
    customer_type_data = round_frame(
        dataframe.groupby("Customer_Type")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"))
        .reset_index()
        .rename(columns={"Customer_Type": "type"})
        .sort_values("sales", ascending=False)
    )
    counts, bin_edges = np.histogram(dataframe["Profit"], bins=10)
    profit_distribution = [
        {
            "range": f"{bin_edges[index]:,.0f} to {bin_edges[index + 1]:,.0f}",
            "count": int(counts[index]),
        }
        for index in range(len(counts))
    ]
    margin_by_category = round_frame(
        dataframe.groupby("Product_Category")
        .agg(avgMargin=("Profit_Margin", "mean"))
        .reset_index()
        .rename(columns={"Product_Category": "category"})
        .sort_values("avgMargin", ascending=False)
    )
    yearly_data = round_frame(
        dataframe.groupby("Year")
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"))
        .reset_index()
        .rename(columns={"Year": "year"})
        .sort_values("year")
    )
    quarterly_data = round_frame(
        dataframe.groupby(["Year", "Quarter"])
        .agg(sales=("Sales_Amount", "sum"), profit=("Profit", "sum"))
        .reset_index()
        .rename(columns={"Year": "year", "Quarter": "quarter"})
        .sort_values(["year", "quarter"])
    )
    forecast_data, forecast_metrics = build_forecast(dataframe)

    return {
        "overallKPIs": {
            "totalSales": round(float(total_sales), 2),
            "totalProfit": round(float(total_profit), 2),
            "totalOrders": int(total_orders),
            "avgOrderValue": round(float(avg_order_value), 2),
            "profitMargin": round(float(profit_margin), 2),
            "totalCustomers": int(dataframe["Customer ID"].nunique()),
        },
        "monthlySales": monthly_sales.to_dict(orient="records"),
        "regionData": region_data.to_dict(orient="records"),
        "categoryData": category_data.to_dict(orient="records"),
        "subCategoryData": subcategory_data.to_dict(orient="records"),
        "topProducts": top_products.to_dict(orient="records"),
        "bottomProducts": bottom_products.to_dict(orient="records"),
        "topCustomers": top_customers.to_dict(orient="records"),
        "topStates": top_states.to_dict(orient="records"),
        "cityData": city_data.to_dict(orient="records"),
        "segmentData": segment_data.to_dict(orient="records"),
        "paymentData": payment_data.to_dict(orient="records"),
        "channelData": channel_data.to_dict(orient="records"),
        "shipModeData": ship_mode_data.to_dict(orient="records"),
        "customerTypeData": customer_type_data.to_dict(orient="records"),
        "profitDistribution": profit_distribution,
        "marginByCategory": margin_by_category.to_dict(orient="records"),
        "forecastData": forecast_data,
        "forecastMetrics": forecast_metrics,
        "yearlyData": yearly_data.to_dict(orient="records"),
        "quarterlyData": quarterly_data.to_dict(orient="records"),
        "generatedAt": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def normalise_excel_value(value):
    """Convert pandas types to Excel-friendly values."""
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, np.generic):
        return value.item()
    return value


def apply_auto_width(worksheet) -> None:
    """Set worksheet column widths based on content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def build_excel_dashboard(dataframe: pd.DataFrame, payload: dict) -> None:
    """Create the Excel workbook with Data, Summary, and Charts_Guide sheets."""
    workbook = Workbook()

    data_sheet = workbook.active
    data_sheet.title = "Data"
    export_frame = dataframe.copy()

    data_sheet.append(list(export_frame.columns))
    for row in export_frame.itertuples(index=False, name=None):
        data_sheet.append([normalise_excel_value(value) for value in row])

    last_row = data_sheet.max_row
    last_column = data_sheet.max_column
    table_ref = f"A1:{get_column_letter(last_column)}{last_row}"
    data_table = Table(displayName="CleanedSalesData", ref=table_ref)
    data_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    data_sheet.add_table(data_table)
    data_sheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor=PRIMARY)
    header_font = Font(color=WHITE, bold=True)
    for cell in data_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    numeric_format_columns = {
        "Sales_Amount",
        "Unit_Price",
        "Unit_Cost",
        "Profit",
        "Profit_Margin",
        "Shipping_Days",
        "Quantity",
    }
    for row in data_sheet.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            header_value = data_sheet.cell(row=1, column=cell.column).value
            if header_value in numeric_format_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet["A1"] = "Sales Performance Dashboard Summary"
    summary_sheet["A1"].font = Font(size=16, bold=True, color=PRIMARY)

    kpis = payload["overallKPIs"]
    kpi_labels = [
        ("Total Sales", kpis["totalSales"]),
        ("Total Profit", kpis["totalProfit"]),
        ("Total Orders", kpis["totalOrders"]),
        ("Avg Order Value", kpis["avgOrderValue"]),
        ("Profit Margin %", kpis["profitMargin"]),
        ("Total Customers", kpis["totalCustomers"]),
    ]
    start_row = 3
    for index, (label, value) in enumerate(kpi_labels, start=0):
        row_num = start_row + index
        summary_sheet[f"A{row_num}"] = label
        summary_sheet[f"B{row_num}"] = value
        summary_sheet[f"A{row_num}"].font = Font(bold=True, color=WHITE)
        summary_sheet[f"A{row_num}"].fill = header_fill
        summary_sheet[f"B{row_num}"].fill = PatternFill(fill_type="solid", fgColor=LIGHT_ROW)
        summary_sheet[f"A{row_num}"].alignment = Alignment(horizontal="left")
        summary_sheet[f"B{row_num}"].alignment = Alignment(horizontal="right")

    summary_sheet["D3"] = "Category Summary"
    summary_sheet["D3"].font = Font(bold=True, color=PRIMARY)
    category_start = 4
    summary_sheet[f"D{category_start}"] = "Category"
    summary_sheet[f"E{category_start}"] = "Sales"
    summary_sheet[f"F{category_start}"] = "Profit"
    for cell in summary_sheet[f"D{category_start}:F{category_start}"][0]:
        cell.fill = header_fill
        cell.font = header_font

    for offset, item in enumerate(payload["categoryData"], start=1):
        summary_sheet[f"D{category_start + offset}"] = item["category"]
        summary_sheet[f"E{category_start + offset}"] = item["sales"]
        summary_sheet[f"F{category_start + offset}"] = item["profit"]

    region_start = 10
    summary_sheet[f"D{region_start}"] = "Region Summary"
    summary_sheet[f"D{region_start}"].font = Font(bold=True, color=PRIMARY)
    summary_sheet[f"D{region_start + 1}"] = "Region"
    summary_sheet[f"E{region_start + 1}"] = "Sales"
    summary_sheet[f"F{region_start + 1}"] = "Profit"
    for cell in summary_sheet[f"D{region_start + 1}:F{region_start + 1}"][0]:
        cell.fill = header_fill
        cell.font = header_font

    for offset, item in enumerate(payload["regionData"], start=2):
        summary_sheet[f"D{region_start + offset}"] = item["region"]
        summary_sheet[f"E{region_start + offset}"] = item["sales"]
        summary_sheet[f"F{region_start + offset}"] = item["profit"]

    charts_guide_sheet = workbook.create_sheet("Charts_Guide")
    charts_guide_sheet["A1"] = "How to Build the Excel Dashboard"
    charts_guide_sheet["A1"].font = Font(size=15, bold=True, color=PRIMARY)
    guide_lines = [
        "1. Create a Pivot Table from the Data sheet table.",
        "2. Build monthly sales trend using Order Date grouped by Months and Years.",
        "3. Create Region, Product_Category, Segment, and Top Product pivots.",
        "4. Add slicers for Region, Segment, Ship Mode, and Customer_Type.",
        "5. Use clustered columns, line charts, donut charts, and KPI cells.",
        "6. Apply the navy and gold color theme for visual consistency.",
        "7. Place KPI cards at the top and charts below in a clean dashboard layout.",
    ]
    for row_index, line in enumerate(guide_lines, start=3):
        charts_guide_sheet[f"A{row_index}"] = line
        charts_guide_sheet[f"A{row_index}"].alignment = Alignment(wrap_text=True)

    for worksheet in workbook.worksheets:
        apply_auto_width(worksheet)

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(EXCEL_FILE)
    print(f"Excel dashboard workbook created: {EXCEL_FILE}")


def main() -> None:
    """Entry point for JSON and Excel export."""
    try:
        configure_console()
        print(f"Reading cleaned dataset: {INPUT_FILE}")
        dataframe = pd.read_csv(INPUT_FILE, parse_dates=["Order Date", "Ship Date"])
        payload = build_dashboard_payload(dataframe)

        DATA_JS_FILE.parent.mkdir(parents=True, exist_ok=True)
        DATA_JS_FILE.write_text(
            "window.dashboardData = " + json.dumps(payload, indent=2) + ";\n",
            encoding="utf-8",
        )
        print(f"Frontend data exported to: {DATA_JS_FILE}")

        build_excel_dashboard(dataframe, payload)
        print("\nStep 06 completed successfully.")
    except Exception as exc:  # pragma: no cover - runtime safety
        print(f"Error in 06_export_json_for_frontend.py: {exc}")
        sys.exit(1)


if __name__ == "__main__":
    main()
